import random
from openpyxl import Workbook, load_workbook

def wellness_question_preprocess():
  input = "./data/웰니스_대화_스크립트_데이터셋.xlsx"
  q_output = "./data/wellness_question.txt"

  f = open(q_output, 'w')

  wb = load_workbook(filename=input)

  ws = wb[wb.sheetnames[0]]
  for row in ws.iter_rows():
    f.write(row[0].value + "    " + row[1].value + "\n")

  f.close()

def wellness_answer_preprocess():
  input = "./data/웰니스_대화_스크립트_데이터셋.xlsx"
  a_output = "./data/wellness_ans.txt"

  f = open(a_output, 'w')
  wb = load_workbook(filename=input)
  ws = wb[wb.sheetnames[0]]

  for row in ws.iter_rows():
    if row[2].value == None:
      continue
    else:
      f.write(row[0].value + "    " + row[2].value + "\n")
  f.close()

def wellness_text_classification_preprocess():
  category_file = "./data/wellness_category.txt"
  question_file = "./data/wellness_question.txt"
  text_classification_file = "./data/wellness_text_classification.txt"

  cate_file = open(category_file, 'r')
  ques_file = open(question_file, 'r')
  text_classfi_file = open(text_classification_file, 'w')

  category_lines = cate_file.readlines()
  cate_dict = {}
  for _, line_data in enumerate(category_lines):
    data = line_data.split('    ')
    cate_dict[data[0]] = data[1][:-1]
  print(cate_dict)

  ques_lines = ques_file.readlines()
  for _, line_data in enumerate(ques_lines):
    data = line_data.split('    ')
    text_classfi_file.write(data[1][:-1] + "    " + cate_dict[data[0]] + "\n")

  cate_file.close()
  ques_file.close()
  text_classfi_file.close()
  
def seperate_wellness_data():
  file_path = "./data/wellness_text_classification.txt"
  train_file_path = "./data/wellness_text_classification_train.txt"
  test_file_path = "./data/wellness_text_classification_test.txt"

  sperated_file = open(file_path, 'r')
  train_file = open(train_file_path, 'w')
  test_file = open(test_file_path, 'w')

  sperated_file_lines = sperated_file.readlines()
  for _, line_data in enumerate(sperated_file_lines):
    rand_num = random.randint(0, 10)
    if rand_num < 10:
      train_file.write(line_data)
    else:
      test_file.write(line_data)

  sperated_file.close()
  train_file.close()
  test_file.close()